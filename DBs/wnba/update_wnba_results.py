#!/usr/bin/env python3
"""
update_wnba_results.py

Fills PF / PA / OT into WNBA_2026_Results.xlsx by pulling final scores from
ESPN's public (unofficial, no key required) scoreboard API, then validates
that every game's two mirror rows (home + away) agree before saving.

USAGE
    python update_wnba_results.py --file WNBA_2026_Results.xlsx --start 2026-05-08 --end 2026-05-27
    python update_wnba_results.py --file WNBA_2026_Results.xlsx --date 2026-05-27
    python update_wnba_results.py --file WNBA_2026_Results.xlsx --start 2026-05-08 --end today

If --start/--end/--date are omitted, it scans the whole sheet and fetches
scores for every date that still has blank PF/PA and is not in the future.

WHAT IT DOES
    1. Reads the sheet, finds rows with blank PF/PA whose Date has passed.
    2. Groups those dates and hits ESPN's scoreboard endpoint once per date.
    3. Matches each returned game to its two mirror rows by (Date, Team, Opp).
    4. Writes PF, PA, OT into both rows.
    5. Runs a validation pass over the WHOLE sheet (not just what it touched):
         - every game's two rows must have mirrored PF/PA (A's PF == H's PA)
         - both rows must have the same OT flag
         - no duplicate (Date, Team, Opp) pairs
         - no past-dated game left blank after the update
       Anything that fails is printed and the row is left untouched — it
       never writes a value it isn't confident about.
    6. Saves the file in place (only if there were no validation failures
       introduced by this run) and prints a short summary.

NOTES ON THE DATA SOURCE
    ESPN's endpoint is unofficial, so team names are matched by substring
    against the TEAM_NAME_MAP below. Two teams are new for 2026 (Toronto,
    Portland) and I could not verify their exact ESPN naming from this
    sandbox (network here is restricted to package registries, not sports
    sites). Run with --diagnose-date on the first real game day to print
    ESPN's raw team names before trusting the mapping, and adjust
    TEAM_NAME_MAP if a team fails to match.
"""

import argparse
import sys
from datetime import date, datetime, timedelta

import openpyxl
import requests

ESPN_SCOREBOARD_URL = "https://site.api.espn.com/apis/site/v2/sports/basketball/wnba/scoreboard"

# Map substrings found in ESPN's team display name -> the code used in the sheet.
# Adjust the TOR / POR entries after running --diagnose-date once real games
# for those franchises are on the schedule.
TEAM_NAME_MAP = {
    "connecticut": "CON",
    "new york": "NYL",
    "golden state": "GSV",
    "seattle": "SEA",
    "washington": "WAS",
    "toronto": "TOR",
    "dallas": "DAL",
    "indiana": "IND",
    "phoenix": "PHX",
    "las vegas": "LVA",
    "atlanta": "ATL",
    "minnesota": "MIN",
    "chicago": "CHI",
    "portland": "POR",
    "los angeles": "LAS",
}


def code_for_espn_team(display_name: str) -> str | None:
    name = display_name.lower()
    for substr, code in TEAM_NAME_MAP.items():
        if substr in name:
            return code
    return None


def fetch_scores_for_date(d: date) -> list[dict]:
    """Return a list of {team_code, opp_code, team_score, opp_score, ot} for
    every FINAL game on this date, one entry per team (i.e. two per game)."""
    resp = requests.get(
        ESPN_SCOREBOARD_URL,
        params={"dates": d.strftime("%Y%m%d")},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()

    out = []
    for event in data.get("events", []):
        status = event.get("status", {}).get("type", {}).get("state")
        if status != "post":  # not finished yet
            continue
        competition = event["competitions"][0]
        competitors = competition["competitors"]
        if len(competitors) != 2:
            continue

        period = competition.get("status", {}).get("period", 4)
        ot = 1.0 if period and period > 4 else 0.0

        resolved = []
        for c in competitors:
            code = code_for_espn_team(c["team"]["displayName"])
            score = c.get("score")
            resolved.append((code, float(score) if score is not None else None))

        (code_a, score_a), (code_b, score_b) = resolved
        if code_a is None or code_b is None:
            print(f"  [!] Could not map a team on {d}: "
                  f"{competitors[0]['team']['displayName']} vs "
                  f"{competitors[1]['team']['displayName']} — skipped, fix TEAM_NAME_MAP")
            continue
        if score_a is None or score_b is None:
            continue

        out.append({"team": code_a, "opp": code_b, "pf": score_a, "pa": score_b, "ot": ot})
        out.append({"team": code_b, "opp": code_a, "pf": score_b, "pa": score_a, "ot": ot})
    return out


def diagnose_date(d: date) -> None:
    resp = requests.get(ESPN_SCOREBOARD_URL, params={"dates": d.strftime("%Y%m%d")}, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    print(f"Raw ESPN team names for {d}:")
    for event in data.get("events", []):
        for c in event["competitions"][0]["competitors"]:
            name = c["team"]["displayName"]
            print(f"  {name!r:35s} -> mapped to {code_for_espn_team(name)}")


def col_index(ws, header_name: str) -> int:
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    raise KeyError(f"Column {header_name!r} not found in header row")


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", required=True, help="Path to WNBA_2026_Results.xlsx")
    ap.add_argument("--start", help="YYYY-MM-DD, or 'today'")
    ap.add_argument("--end", help="YYYY-MM-DD, or 'today'")
    ap.add_argument("--date", help="Single date YYYY-MM-DD, shorthand for --start/--end the same day")
    ap.add_argument("--diagnose-date", help="Print raw ESPN team names for YYYY-MM-DD and exit, no file changes")
    args = ap.parse_args()

    if args.diagnose_date:
        diagnose_date(datetime.strptime(args.diagnose_date, "%Y-%m-%d").date())
        return

    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    c_date = col_index(ws, "Date")
    c_team = col_index(ws, "Team")
    c_opp = col_index(ws, "Opp")
    c_pf = col_index(ws, "PF")
    c_pa = col_index(ws, "PA")
    c_ot = col_index(ws, "OT")

    rows = list(ws.iter_rows(min_row=2))

    # Figure out which dates to fetch.
    today = date.today()
    if args.date:
        target_dates = {datetime.strptime(args.date, "%Y-%m-%d").date()}
    elif args.start or args.end:
        start = today if args.start in (None, "today") else datetime.strptime(args.start, "%Y-%m-%d").date()
        end = today if args.end in (None, "today") else datetime.strptime(args.end, "%Y-%m-%d").date()
        target_dates = {start + timedelta(days=i) for i in range((end - start).days + 1)}
    else:
        # Every past date that still has a blank score somewhere.
        target_dates = set()
        for r in rows:
            dcell = r[c_date - 1].value
            pfcell = r[c_pf - 1].value
            if dcell is None:
                continue
            gdate = dcell.date() if hasattr(dcell, "date") else dcell
            if pfcell is None and gdate <= today:
                target_dates.add(gdate)

    target_dates = sorted(d for d in target_dates if d <= today)
    if not target_dates:
        print("No past/blank dates to update.")
    else:
        print(f"Fetching {len(target_dates)} date(s): {target_dates[0]} to {target_dates[-1]}")

    # Build a lookup of (date, team, opp) -> row, restricted to blank rows.
    lookup = {}
    for r in rows:
        dcell = r[c_date - 1].value
        if dcell is None:
            continue
        gdate = dcell.date() if hasattr(dcell, "date") else dcell
        key = (gdate, r[c_team - 1].value, r[c_opp - 1].value)
        lookup[key] = r

    filled, already_had_value, unmatched = 0, 0, 0

    for d in target_dates:
        try:
            games = fetch_scores_for_date(d)
        except requests.RequestException as e:
            print(f"  [!] Failed to fetch {d}: {e}")
            continue

        for g in games:
            key = (d, g["team"], g["opp"])
            row = lookup.get(key)
            if row is None:
                unmatched += 1
                print(f"  [!] No matching schedule row for {d} {g['team']} vs {g['opp']} — game not on sheet?")
                continue
            existing_pf = row[c_pf - 1].value
            if existing_pf is not None:
                already_had_value += 1
                continue
            row[c_pf - 1].value = g["pf"]
            row[c_pa - 1].value = g["pa"]
            row[c_ot - 1].value = g["ot"]
            filled += 1

    print(f"Filled {filled} row(s). {already_had_value} already had scores. {unmatched} unmatched game(s).")

    # ---- Validation pass over the whole sheet before saving ----
    errors = []
    seen_pairs = {}
    for r in rows:
        dcell = r[c_date - 1].value
        if dcell is None:
            continue
        gdate = dcell.date() if hasattr(dcell, "date") else dcell
        team, opp = r[c_team - 1].value, r[c_opp - 1].value
        pf, pa, ot = r[c_pf - 1].value, r[c_pa - 1].value, r[c_ot - 1].value

        if pf is None and gdate < today:
            errors.append(f"{gdate} {team} vs {opp}: game is in the past but still has no score")
            continue
        if pf is None:
            continue  # future game, fine

        mirror_key = (gdate, opp, team)  # the other team's row
        mirror = lookup.get(mirror_key)
        if mirror is None:
            errors.append(f"{gdate} {team} vs {opp}: no mirror row found at all")
            continue
        m_pf, m_pa, m_ot = mirror[c_pf - 1].value, mirror[c_pa - 1].value, mirror[c_ot - 1].value
        if m_pf is None:
            errors.append(f"{gdate} {team} vs {opp}: mirror row ({opp} vs {team}) has no score")
            continue
        if pf != m_pa or pa != m_pf:
            errors.append(f"{gdate} {team}({pf}-{pa}) vs mirror {opp}({m_pf}-{m_pa}): scores don't mirror")
        if ot != m_ot:
            errors.append(f"{gdate} {team} vs {opp}: OT flag mismatch ({ot} vs {m_ot})")

        dupe_key = (gdate, team, opp)
        if dupe_key in seen_pairs:
            errors.append(f"{gdate} {team} vs {opp}: duplicate row")
        seen_pairs[dupe_key] = True

    if errors:
        print(f"\n{len(errors)} validation issue(s) found — NOT saving the file:")
        for e in errors:
            print(f"  [!] {e}")
        sys.exit(1)

    wb.save(args.file)
    print(f"\nSaved {args.file} — sheet passed validation.")


if __name__ == "__main__":
    main()
